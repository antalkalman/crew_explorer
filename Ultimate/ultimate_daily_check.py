import os
import re
import glob
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# === Setup paths ===
desktop_path = "/Users/kalmanantal/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive"
laptop_path = "/Users/antalkalman/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive"
base_dir = desktop_path if os.path.exists(desktop_path) else laptop_path
temp_dir = os.path.join(base_dir, "Temp")
os.makedirs(temp_dir, exist_ok=True)

# === Step 1: Run external scripts ===
os.system("python3 ultimate_two_sflist_state.py")
os.system("python3 ultimate_compare_two_sflists.py")

# === Step 2: Load latest CSVs from Temp ===
def find_latest_csv(prefix):
    files = glob.glob(os.path.join(temp_dir, f"{prefix}*.csv"))
    if not files:
        return None, None
    def extract_timestamp(file):
        match = re.search(r"_(\d{8}_\d{4})", os.path.basename(file))
        return datetime.strptime(match.group(1), "%Y%m%d_%H%M") if match else datetime.min
    latest_file = max(files, key=extract_timestamp)
    df = pd.read_csv(latest_file)
    return df, os.path.basename(latest_file)

df_sf_issues, fn1 = find_latest_csv("SF_Issues_")
df_new_issues, fn2 = find_latest_csv("New_Issues_")
df_resolved_issues, fn3 = find_latest_csv("Resolved_Issues_")
df_cmp, fn4 = find_latest_csv("SFlist_changes_")

# === Step 3: Load latest SFlist_*.csv from base_dir ===
sflist_files = glob.glob(os.path.join(base_dir, "SFlist_*.csv"))
latest_sflist = max(sflist_files, key=os.path.getmtime)
df_sflist = pd.read_csv(latest_sflist)

# === Step 4: Export all to Excel ===
timestamp = datetime.now().strftime("%Y%m%d_%H%M")
output_excel = os.path.join(base_dir, f"Full_SF_Overview_{timestamp}.xlsx")

def extract_timestamp(filename):
    match = re.search(r"(\d{8}_\d{4})", filename)
    return match.group(1) if match else timestamp

with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
    for df, name, base_tab in [
        (df_sf_issues, fn1, "SF Issues"),
        (df_new_issues, fn2, "New Issues"),
        (df_resolved_issues, fn3, "Resolved Issues"),
        (df_cmp.assign(**{
            "Change Status": df_cmp["Change Status"].apply(
                lambda x: ", ".join(x.split(", ")[:3]) + ("..." if len(x.split(", ")) > 3 else "")
                if isinstance(x, str) else x)
        }), fn4, "SFlist Changes"),
        (df_sflist, os.path.basename(latest_sflist), "Latest SFlist")
    ]:
        if df is not None:
            ts = extract_timestamp(name)
            tab_name = f"{base_tab} {ts}"[:31]
            df.to_excel(writer, sheet_name=tab_name, index=False)

            ws = writer.sheets[tab_name]
            end_col = get_column_letter(ws.max_column)
            end_row = ws.max_row
            table_range = f"A1:{end_col}{end_row}"

            table_name = re.sub(r'\W+', '', tab_name)[:31]
            table = Table(displayName=table_name, ref=table_range)
            table.tableStyleInfo = TableStyleInfo(
                name=None, showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False
            )
            ws.add_table(table)
            #tet


            # Auto-adjust other columns, skipping "Change Status"
            for col in ws.columns:
                header = col[0].value
                if header == "Change Status":
                    continue
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 2

            # Set narrower width for "Change Status"
            for col_cells in ws.iter_cols(min_row=1, max_row=1):
                for cell in col_cells:
                    if cell.value == "Change Status":
                        col_letter = cell.column_letter
                        ws.column_dimensions[col_letter].width = 25


print(f"✅ Full report saved to: {output_excel}")
