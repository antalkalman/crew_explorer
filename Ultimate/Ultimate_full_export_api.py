import os
import time
import requests
import pandas as pd
import ast
import csv
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter



# === CONFIG ===

# Automatically detect base directory
desktop_path = "/Users/kalmanantal/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive"
laptop_path = "/Users/antalkalman/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive"

base_dir = desktop_path if os.path.exists(desktop_path) else laptop_path

# Token and API
token = "4|FtRaHbpkyP7dj6geLbqQiUc2WznEEqbRhvOD2xPO1cee3bfd"
base_url = "https://manager.crewcall.hu/api/"
headers = {"Authorization": f"Bearer {token}"}

# File paths
mapping_file = os.path.join(base_dir, "export_field_mapping.xlsx")
output_file = os.path.join(base_dir, "Formatted_SFlist_Mapped.csv")

# List of project names to include
included_projects = ["3BP", "Oasis", "Seven Sisters", "BETAMAX", "Hidden Hand 2"]
#included_projects = ["FBI S4", "Ballerina Overdrive", "Blue Moon"]
#included_projects = ["Test"]

# Make sure folder exists
os.makedirs(base_dir, exist_ok=True)

# === FUNCTION: Fetch data from Crew Manager API ===
def fetch_data(endpoint, retries=3, delay=5):
    for attempt in range(retries):
        try:
            response = requests.get(
                url=f"{base_url}{endpoint}",
                headers=headers,
                timeout=30
            )
            response.raise_for_status()
            return response.json().get("data", [])
        except requests.exceptions.RequestException as e:
            print(f"❌ {endpoint} – Attempt {attempt + 1} failed: {e}")
            if attempt < retries - 1:
                time.sleep(delay)
            else:
                raise
    return None

# === FETCH ALL API DATA ===
projects = fetch_data("project")
departments = fetch_data("department")
job_titles = fetch_data("job_title")
startforms = fetch_data("startform")
users = fetch_data("user")
overtimes = fetch_data("overtime")
templates = fetch_data("startform_template")
turnarounds = fetch_data("turnaround")
units = fetch_data("unit")
working_hours = fetch_data("working_hour")

# === LOOKUP DICTIONARIES ===
project_lookup = {p["id"]: p.get("name", "") for p in projects}
department_lookup = {d["id"]: d.get("name", "") for d in departments}
job_title_lookup = {j["id"]: j.get("name", "") for j in job_titles}
department_sort_lookup = {d["id"]: d.get("sort", 0) for d in departments}
job_title_sort_lookup = {j["id"]: j.get("sort", 0) for j in job_titles}
overtime_lookup = {o["id"]: o.get("name", "") for o in overtimes}
turnaround_lookup = {t["id"]: t.get("name", "") for t in turnarounds}
unit_lookup = {u["id"]: u.get("name", "") for u in units}
working_hour_lookup = {w["id"]: w.get("name", "") for w in working_hours}
template_lookup = {t["id"]: t.get("title", "") for t in templates}
unit_sort_lookup = {u["id"]: u.get("sort", 0) for u in units}




# === LOAD AND NORMALIZE SF DATA ===
df_mapping = pd.read_excel(mapping_file)
df_sf = pd.json_normalize(startforms)

# === USER LOOKUP EXPANSION ===
user_lookup = {
    u["id"]: {
        "User name": u.get("name", ""),
        "User surname": u.get("surname", ""),
        "User email": u.get("email", ""),
        "User phone": u.get("phone", "")
    }
    for u in users
}

user_df = df_sf["user_id"].map(user_lookup).apply(pd.Series)
df_sf = pd.concat([df_sf, user_df], axis=1)


# === DATE CLEANUP ===
# Normalize date columns to be tz-naive (remove timezone info)
for col in df_sf.columns:
    try:
        if isinstance(col, str) and "date" in col.lower():
            df_sf[col] = pd.to_datetime(df_sf[col], errors="coerce").dt.tz_localize(None)
    except Exception as e:
        print(f"⚠️ Could not normalize column '{col}': {e}")


# === MAP LOOKUPS ===
df_sf["Project"] = df_sf["project_id"].map(project_lookup)
df_sf["Project department"] = df_sf["project_department_id"].map(department_lookup)
df_sf["Project job title"] = df_sf["project_job_title_id"].map(job_title_lookup)
df_sf["Project overtime"] = df_sf["project_overtime_id"].map(overtime_lookup)
df_sf["Project turnaround"] = df_sf["project_turnaround_id"].map(turnaround_lookup)
df_sf["Project unit"] = df_sf["project_unit_id"].map(unit_lookup)
df_sf["Project working hour"] = df_sf["project_working_hour_id"].map(working_hour_lookup)
df_sf["Startform template"] = df_sf["project_startform_id"].map(template_lookup)
df_sf["Unit Sort"] = df_sf["project_unit_id"].map(unit_sort_lookup)




# === FILTER BY PROJECT ===
df_sf = df_sf[df_sf["Project"].isin(included_projects)]


# Clean up deal_notes
def clean_deal_note(val):
    if isinstance(val, str):
        return val.replace("\n", " ").replace("\r", " ").replace('"', '').replace(",", ";").strip()
    return val

if "deal_notes" in df_sf.columns:
    df_sf["deal_notes"] = df_sf["deal_notes"].apply(clean_deal_note)

# Derived fields
df_sf["Sf number"] = df_sf["type"].fillna("") + df_sf["sf_number"].astype(str)
df_sf["Crew member id"] = df_sf["crew_member_id"].apply(lambda x: f"CM{int(x)}" if pd.notnull(x) else "")

# New derived sort fields
df_sf["Dept Sort"] = df_sf["project_department_id"].map(department_sort_lookup)
df_sf["Title Sort"] = df_sf["project_job_title_id"].map(job_title_sort_lookup)

# === FLATTEN nested `others` fields ===
def extract_others(entry_list, label):
    result = {}
    try:
        if isinstance(entry_list, list):
            for i in range(min(2, len(entry_list))):
                item = entry_list[i] or {}
                result[f"{label} {i+1} description"] = str(item.get("name", "") or "")
                result[f"{label} {i+1} price"] = str(item.get("price", "") or "")
                result[f"{label} {i+1} account code"] = str(item.get("account_code", "") or "")
        else:
            parsed = ast.literal_eval(entry_list) if isinstance(entry_list, str) else []
            return extract_others(parsed, label)
    except Exception as e:
        result[f"{label} 1 description"] = f"⚠️ error: {e}"
    return result

# Apply flattening
df_daily = df_sf["daily_others"].apply(lambda x: pd.Series(extract_others(x, "Daily others")))
df_weekly = df_sf["weekly_others"].apply(lambda x: pd.Series(extract_others(x, "Weekly others")))
df_fee = df_sf["fee_others"].apply(lambda x: pd.Series(extract_others(x, "Fee others")))

# Combine all flattened columns
df_sf = pd.concat([df_sf, df_daily, df_weekly, df_fee], axis=1)

# Normalize date columns to be tz-naive (remove timezone info)
for col in df_sf.columns:
    if isinstance(col, str) and "date" in col.lower():
        try:
            df_sf[col] = pd.to_datetime(df_sf[col], errors="coerce").dt.tz_localize(None)
        except Exception as e:
            print(f"⚠️ Could not normalize column '{col}': {e}")


# === BUILD FINAL OUTPUT ===
output_dict = {}

for _, row in df_mapping.iterrows():
    export_col = row["Export Column (Official)"]
    api_field = row["API Field (from CrewManager_StartForms.csv)"]

    if export_col in df_sf.columns and (pd.isna(api_field) or api_field.strip() == ""):
        output_dict[export_col] = df_sf[export_col]
    elif pd.isna(api_field) or api_field.strip() == "":
        output_dict[export_col] = ""
    elif " + " in api_field:
        try:
            parts = [df_sf[p.strip()] if p.strip() in df_sf.columns else "" for p in api_field.split("+")]
            output_dict[export_col] = parts[0].astype(str) + parts[1].astype(str)
        except Exception:
            output_dict[export_col] = ""
    elif api_field.startswith('"CM" +'):
        field = api_field.split("+")[1].strip()
        output_dict[export_col] = df_sf[field].apply(lambda x: f"CM{int(x)}" if pd.notnull(x) else "")
    elif api_field in df_sf.columns:
        output_dict[export_col] = df_sf[api_field]
    else:
        output_dict[export_col] = ""

output_data = pd.DataFrame(output_dict)

# Move specific columns to the end
cols_to_move = [
    "is_internal", "invite_date", "sort_order",
    "created_at", "updated_at", "deleted_at", "downloaded_at"
]

# Keep all other columns in original order
final_columns = [col for col in output_data.columns if col not in cols_to_move] + cols_to_move
output_data = output_data[final_columns]


# === EXPORT CSV WITH TIMESTAMP ===
timestamp = datetime.now().strftime("%Y%m%d_%H%M")
output_file = os.path.join(base_dir, f"SFlist_{timestamp}.csv")

output_data.to_csv(output_file, index=False, quoting=csv.QUOTE_ALL)
print(f"✅ Final formatted CSV saved to:\n{output_file}")

"""
# === OPTIONAL: EXPORT TO EXCEL WITH TABLE ===
# Uncomment this block to enable Excel export
excel_output_file = output_file.replace(".csv", ".xlsx")
output_data.to_excel(excel_output_file, index=False)

# Add Excel Table formatting
from openpyxl.utils import get_column_letter

wb = load_workbook(excel_output_file)
ws = wb.active
last_col_letter = get_column_letter(len(output_data.columns))
table_ref = f"A1:{last_col_letter}{len(output_data) + 1}"
table = Table(displayName="CrewExport", ref=table_ref)
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
table.tableStyleInfo = style
ws.add_table(table)
wb.save(excel_output_file)
print(f"📘 Excel version with table saved to:\n{excel_output_file}")

"""
