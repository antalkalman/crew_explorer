import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# === Set folder path for both machines ===
desktop_path = "/Users/kalmanantal/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive"
laptop_path  = "/Users/antalkalman/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive"
compare_dir  = desktop_path if os.path.exists(desktop_path) else laptop_path


# === Match "Checked_StartForms_based_on_*.xlsx" files ===
pattern = re.compile(r"Checked_StartForms_based_on_.*_created_(\d{8}_\d{4})\.xlsx$")
files = []

for f in os.listdir(compare_dir):
    match = pattern.match(f)
    if match:
        try:
            ts = datetime.strptime(match.group(1), "%Y%m%d_%H%M")
            files.append((ts, f))
        except ValueError:
            continue

if len(files) < 2:
    raise FileNotFoundError("❌ Need at least two Checked_StartForms_based_on_*.xlsx files in Compare_Issues folder.")

# Sort files descending by timestamp
files.sort(reverse=True)
newer_file = files[0][1]
older_file = files[1][1]

print(f"🔍 Comparing:\n  NEW : {newer_file}\n  OLD : {older_file}")

# === Load data ===
new_book = pd.read_excel(os.path.join(compare_dir, newer_file), sheet_name=None)
old_book = pd.read_excel(os.path.join(compare_dir, older_file), sheet_name=None)

df_new = new_book["PM View"] if "PM View" in new_book else get_first_valid_sheet(new_book)
df_old = old_book["PM View"] if "PM View" in old_book else get_first_valid_sheet(old_book)



# Load from correct sheet (PM View if exists, fallback to first sheet)
def get_first_valid_sheet(sheet_dict):
    for key in sheet_dict:
        if isinstance(sheet_dict[key], pd.DataFrame):
            return sheet_dict[key]
    raise ValueError("❌ No valid sheets found.")


# Drop completely empty rows
df_new = df_new.dropna(how="all")
df_old = df_old.dropna(how="all")

# === Normalize for matching ===
df_new["SF Key"] = df_new["SF Key"].astype(str).str.strip().str.lower()
df_old["SF Key"] = df_old["SF Key"].astype(str).str.strip().str.lower()

# === Identify New Issues ===
old_keys = set(df_old["SF Key"])
new_keys = set(df_new["SF Key"])

new_only = df_new[~df_new["SF Key"].isin(old_keys)].copy()
new_only.insert(0, "Status", "New")

# === Identify Resolved Issues ===
# Keep rows from OLD where Issues is NOT empty AND now in NEW Issues is empty
df_old_issues = df_old[df_old["Issues"].astype(str).str.strip() != ""].copy()
resolved = df_old_issues[~df_old_issues["SF Key"].isin(df_new["SF Key"])].copy()

# Find where the issue is gone
resolved = df_old_issues[~df_old_issues["SF Key"].isin(df_new["SF Key"])].copy()


# === Save result ===
older_timestamp = files[1][0].strftime("%Y%m%d_%H%M")
newer_timestamp = files[0][0].strftime("%Y%m%d_%H%M")
output_filename = f"Resolved_Comparison_{older_timestamp}_to_{newer_timestamp}.xlsx"
output_path = os.path.join(compare_dir, output_filename)


with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    new_only.to_excel(writer, sheet_name="New Issues", index=False)
    resolved.to_excel(writer, sheet_name="Resolved Issues", index=False)
    # Also include raw PM View data from both files
    df_new.to_excel(writer, sheet_name="New File", index=False)
    df_old.to_excel(writer, sheet_name="Old File", index=False)


# === Apply Excel Table formatting ===
wb = load_workbook(output_path)

sheets_to_format = ["New Issues", "Resolved Issues", "New File", "Old File"]

for i, sheet_name in enumerate(sheets_to_format, start=1):

    ws = wb[sheet_name]

    if ws.max_row <= 1:
        print(f"⚠️ Skipping table creation for empty sheet: {sheet_name}")
        continue  # Don't apply table formatting to empty sheets

    end_col = get_column_letter(ws.max_column)
    end_row = ws.max_row
    table_range = f"A1:{end_col}{end_row}"

    table_name = f"Table{i}"  # Unique table names: Table1, Table2, etc.
    table = Table(displayName=table_name, ref=table_range)
    style = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2


wb.save(output_path)
print(f"✅ Excel comparison saved to:\n📁 {output_path}")
