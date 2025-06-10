import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# === Set folder where files to compare are placed ===
desktop_compare = "/Users/kalmanantal/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive/Compare"
laptop_compare = "/Users/antalkalman/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive/Compare"
compare_dir = desktop_compare if os.path.exists(desktop_compare) else laptop_compare

# === Match SFlist_YYYYMMDD_HHMM.csv filenames ===
pattern = re.compile(r"SFlist_(\d{8}_\d{4})\.csv$")
sf_files = []

for f in os.listdir(compare_dir):
    match = pattern.match(f)
    if match:
        try:
            ts = datetime.strptime(match.group(1), "%Y%m%d_%H%M")
            sf_files.append((ts, f))
        except ValueError:
            continue

if len(sf_files) < 2:
    raise FileNotFoundError("❌ Need at least two SFlist_YYYYMMDD_HHMM.csv files in the 'Compare' folder.")

# Sort descending by timestamp
sf_files.sort(reverse=True)
latest_file = sf_files[0][1]
older_file = sf_files[1][1]

print(f"🔍 Comparing:\n  NEW : {latest_file}\n  OLD : {older_file}")

# === Load CSVs ===
sf_list = pd.read_csv(os.path.join(compare_dir, latest_file), dtype=str).fillna("")
af_list_old = pd.read_csv(os.path.join(compare_dir, older_file), dtype=str).fillna("")

# === Load compare_directions.xlsx ===
directions_path = os.path.join(compare_dir, "compare_directions.xlsx")
df_directions = pd.read_excel(directions_path)
df_directions.columns = [col.strip() for col in df_directions.columns[:2]]
field_to_category = dict(zip(df_directions.iloc[:, 0].astype(str).str.strip(), df_directions.iloc[:, 1].astype(str).str.strip()))

# === Normalize ===
def normalize(df):
    return df.applymap(lambda x: x.strip().lower() if isinstance(x, str) else x)

sf_clean = normalize(sf_list)
af_clean = normalize(af_list_old)
af_dict = af_clean.set_index("ID").to_dict(orient="index")

# === Compare rows ===
def compare_row(row):
    row_id = row["ID"]
    if row_id not in af_dict:
        return "new"
    old_row = af_dict[row_id]
    diffs = []
    for col in sf_clean.columns:
        category = field_to_category.get(col, "")
        if col == "ID" or category in {"skip", "del"}:
            continue
        if row[col] != old_row.get(col, ""):
            diffs.append(col)
    return ", ".join(diffs) if diffs else "unchanged"

sf_clean["Change Status"] = sf_clean.apply(compare_row, axis=1)

# === Merge results with original data ===
output = sf_list.copy()
output["Change Status"] = sf_clean["Change Status"]

# === Add Category column ===
def get_primary_category(change_str):
    if change_str in ("", "unchanged", "new"):
        return ""
    first_field = change_str.split(",")[0].strip()
    return field_to_category.get(first_field, "")

output["Category"] = output["Change Status"].apply(get_primary_category)

# Force consistent Category values for unchanged and new rows
output.loc[output["Change Status"] == "unchanged", "Category"] = "unchanged"
output.loc[output["Change Status"] == "new", "Category"] = "new"



# === Drop fields marked as 'del' ===
del_fields = [field for field, cat in field_to_category.items() if cat == "del"]
columns_to_keep = [col for col in output.columns if col not in del_fields]
output = output[columns_to_keep]

# === Final Excel export only ===
timestamp = datetime.now().strftime("%Y%m%d_%H%M")
excel_filename = f"SFlist_comparison_output_{timestamp}.xlsx"
excel_path = os.path.join(compare_dir, excel_filename)

# === Reorder columns: Category first, then Change Status ===
cols = output.columns.tolist()
cols.remove("Category")
cols.remove("Change Status")
output = output[["Category", "Change Status"] + cols]


with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    output.to_excel(writer, sheet_name="Comparison", index=False)
    sf_list.to_excel(writer, sheet_name="New SFlist", index=False)
    af_list_old.to_excel(writer, sheet_name="Old SFlist", index=False)

# === Format 'Comparison' sheet as Excel Table ===
wb = load_workbook(excel_path)
ws = wb["Comparison"]
end_col = get_column_letter(ws.max_column)
end_row = ws.max_row
table_range = f"A1:{end_col}{end_row}"

table = Table(displayName="ComparisonResult", ref=table_range)
style = TableStyleInfo(
    name="TableStyleLight1",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=False,
    showColumnStripes=False
)
table.tableStyleInfo = style
ws.add_table(table)

"""
# Optional: Adjust column widths
for col in ws.columns:
    max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_len + 2
"""

wb.save(excel_path)
print(f"📘 Excel with full details saved to:\n{excel_path}")
