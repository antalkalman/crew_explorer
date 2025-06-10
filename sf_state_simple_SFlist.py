import pandas as pd
import numpy as np
import re
from datetime import datetime
import os

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# === Point to SF_Archive folder used by full_export_api_SFlist.py ===
desktop_path = "/Users/kalmanantal/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive"
laptop_path = "/Users/antalkalman/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive"
base_dir = desktop_path if os.path.exists(desktop_path) else laptop_path

# Pattern to match: SFlist_YYYYMMDD_HHMM.csv
pattern = re.compile(r"SFlist_(\d{8}_\d{4})\.csv$")

# === Find the latest SFlist_YYYYMMDD_HHMM.csv in base_dir ===
import os
import re
from datetime import datetime

desktop_path = "/Users/kalmanantal/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive"
laptop_path = "/Users/antalkalman/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive"
base_dir = desktop_path if os.path.exists(desktop_path) else laptop_path

pattern = re.compile(r"SFlist_(\d{8}_\d{4})\.csv$")

latest_file = None
latest_time = None
latest_timestamp_str = None

for file in os.listdir(base_dir):
    match = pattern.match(file)
    if match:
        timestamp_str = match.group(1)
        try:
            file_time = datetime.strptime(timestamp_str, "%Y%m%d_%H%M")
            if not latest_time or file_time > latest_time:
                latest_time = file_time
                latest_file = file
                latest_timestamp_str = timestamp_str  # save valid timestamp
        except ValueError:
            continue

if not latest_file:
    raise FileNotFoundError("❌ No valid SFlist_YYYYMMDD_HHMM.csv file found.")

path_new = os.path.join(base_dir, latest_file)

now_str = datetime.now().strftime("%Y%m%d_%H%M")
output_path = os.path.join(
    base_dir,
    f"Checked_StartForms_based_on_{latest_timestamp_str}_created_{now_str}.xlsx"
)


excluded_projects = {"Death By Lightning", "Test", "Ballerina Overdrive", "Icebreaker"}

# === Load file ===
df = pd.read_csv(path_new)

# Exclude irrelevant projects and paused / crew list SFs
df = df[~df["Project"].isin(excluded_projects)]
df = df[df["State"] != "paused"]
df = df[~df["Sf number"].astype(str).str.upper().str.startswith("CL")]

# === Keep selected columns ===
columns_to_keep = [
    "ID", "Sf number", "Crew member id", "Project", "Currency", "User name", "User surname",
    "User email", "User phone", "Project department", "Project job title", "Project unit", "Title note", "State",
    "Surname", "Firstname", "Nickname", "Email", "Mobile number", "Crew list name", "Crew email", "Citizenship",
    "Start date", "End date", "Deal type",
    "Personal: Tax number / Adóazonosító jel", "Personal: Bank account number / Bankszámlaszám",
    "Company: VAT number / Adószám", "Company: Company name / Cégnév", "Company: Bank account number / Bankszámlaszám",
    "Daily fee", "Car allowance", "Phone allowance", "Computer allowance", "Offset meal",
    "Daily others 1 price", "Daily others 2 price",
    "Weekly fee", "Box rental", "Weekly others price", "Max computer allowance", "Fee others 1 price",
    "Project overtime", "Project turnaround", "Project working hour",
    "Bank account number"
]
df = df[columns_to_keep].copy()

# === Helper functions ===
def safe_str(x):
    return str(x).strip() if not pd.isna(x) else ""

def clean_account(raw):
    digits = re.sub(r'\D', '', str(raw))
    return f"{digits[:8]}-{digits[8:16]}-{digits[16:24]}" if len(digits) >= 16 else digits

def merge_tax_number(df):
    df["Personal: Tax number / Adóazonosító jel"] = df["Personal: Tax number / Adóazonosító jel"].fillna("").astype(str).str.strip()
    df["Company: VAT number / Adószám"] = df["Company: VAT number / Adószám"].fillna("").astype(str).str.strip()
    df["Tax Number"] = df["Personal: Tax number / Adóazonosító jel"]
    df["Tax Number"] = df["Tax Number"].where(df["Tax Number"] != "", df["Company: VAT number / Adószám"])
    df["Tax Number"] = df["Tax Number"].replace("", np.nan)
    return df

def merge_bank_account(df):
    df["Bank Account"] = df["Personal: Bank account number / Bankszámlaszám"].apply(safe_str)
    df["Bank Account"] = df["Bank Account"].where(df["Bank Account"] != "", df["Company: Bank account number / Bankszámlaszám"].apply(safe_str))
    df["Bank Account"] = df["Bank Account"].where(df["Bank Account"] != "", df["Bank account number"].apply(safe_str))
    df["Bank Account"] = df["Bank Account"].apply(clean_account)
    return df

def is_effectively_blank(s):
    if pd.isna(s):
        return True
    s = str(s).strip()
    return s == "" or s.lower() == "nan" or not re.search(r'[A-Za-z0-9]', s)

def has_fee(row):
    return not (pd.isna(row.get("Daily fee")) and pd.isna(row.get("Weekly fee")))

def find_issues(row, today):
    issues = []
    sf_type = str(row.get("Sf number", ""))[:2]

    if row.get("State") not in {"accepted", "signed"} and pd.notna(row.get("Start date")) and row["Start date"] < today:
        days_late = (today - row["Start date"]).days
        issues.append(f"Start date in past but not yet signed or accepted ({days_late} days)")

    if row.get("State") in {"accepted", "signed"}:
        required = [
            "Project department", "Project job title", "Project unit",
            "Surname", "Firstname", "Mobile number",
            "Crew list name", "Crew email", "Start date", "End date", "Deal type",
            "Project overtime", "Project turnaround", "Project working hour"
        ]

        if sf_type == "BD":
            required = [f for f in required if f not in {"Surname", "Firstname", "Mobile number", "Crew list name", "Crew email"}]

        for field in required:
            if is_effectively_blank(row.get(field, "")):
                issues.append(field)

        if not has_fee(row):
            issues.append("Daily fee / Weekly fee")

        tax_number = row.get("Tax Number", "")
        company_name = row.get("Company: Company name / Cégnév", "")
        bank_account = row.get("Bank Account", "")

        if sf_type in {"BD", "DL"}:
            if is_effectively_blank(tax_number) and is_effectively_blank(company_name):
                issues.append("Tax Number/Company")
        else:
            if is_effectively_blank(tax_number):
                issues.append("Tax Number")
            if is_effectively_blank(bank_account):
                issues.append("Bank Account")

    return ", ".join(issues)

# === Preprocessing ===
df["Start date"] = pd.to_datetime(df["Start date"], errors="coerce")
today = pd.Timestamp(datetime.today().date())

df = merge_tax_number(merge_bank_account(df))
df["SF Key"] = df["Sf number"].astype(str) + " - " + df["Project"].astype(str)
df["Issues"] = df.apply(lambda x: find_issues(x, today), axis=1)

# === Responsible and Priority ===
account_keywords = {"Tax Number", "Bank Account"}
pm_keywords = {
    "Start date", "End date",
    "Daily fee", "Weekly fee",
    "Project overtime", "Project turnaround", "Project working hour",
    "Project unit", "Deal type"
}


def determine_responsible(issues):
    parts = [x.strip() for x in issues.split(",") if x.strip()]
    is_pm = any(any(k in part for k in pm_keywords) for part in parts)
    is_account = any(any(k in part for k in account_keywords) for part in parts)
    if is_pm and is_account:
        return "PM, Account"
    elif is_pm:
        return "PM"
    elif is_account:
        return "Account"
    return "Admin"


def determine_priority(issues):
    parts = [x.strip() for x in issues.split(",") if x.strip()]
    if any(part in ["Start date", "End date", "Daily fee / Weekly fee", "Project overtime", "Project turnaround", "Project working hour"] for part in parts):
        return 1
    return 3

df_issues = df[df["Issues"] != ""].copy()
df_issues["Responsible"] = df_issues["Issues"].apply(determine_responsible)
df_issues["Priority"] = df_issues["Issues"].apply(determine_priority)

def determine_category(row):
    if "Start date in past but not yet signed or accepted" in row["Issues"]:
        state = row["State"].strip().lower()
        if state in {"sent", "in progress", "rejected"}:
            return "Chase"
        elif state == "draft":
            return "Plan?"
        elif state == "reviewing":
            return "Accept please"
    return row["Responsible"]

df_issues["Category"] = df_issues.apply(determine_category, axis=1)

# Assign "Who" based on Category
category_to_who = {
    "Chase": "Admin",
    "Accept please": "PM",
    "Plan?": "PM",
    "PM": "PM",
    "PM, Account": "PM, Account",
    "Account": "Account",
    "Admin": "Admin"
}

df_issues["Who"] = df_issues["Category"].map(category_to_who)



# === Sorting ===
df_issues["ResponsibleOrder"] = df_issues["Responsible"].map({"PM": 1, "PM, Account": 2, "Account": 3})
df_issues = df_issues.sort_values(by=["Project", "Priority", "ResponsibleOrder", "Sf number"])
df_issues = df_issues.drop(columns=["ResponsibleOrder"])

# === Define PM View columns ===
pm_columns = [
    "Project",
    "SF Key",
    "Sf number",
    "State",
    "Crew list name",
    "Project department",
    "Project job title",
    "Category",
    "Who",
    "Issues",
    "Priority"
]

df_pm_view = df_issues[pm_columns].copy()

# Define category sort order
category_order = {
    "Chase": 1,
    "Plan?": 2,
    "PM": 3,
    "PM, Account": 4,
    "Account": 5,
    "Accept please": 6
}
df_pm_view["CategorySort"] = df_pm_view["Category"].map(category_order)

# Extract delay days from "Issues"
def extract_days(issue_text):
    match = re.search(r"\((\d+)\s+days\)", issue_text)
    return int(match.group(1)) if match else 0

df_pm_view["DelayDays"] = df_pm_view["Issues"].apply(extract_days)

# Final sort
df_pm_view = df_pm_view.sort_values(by=["Project", "CategorySort", "DelayDays"], ascending=[True, True, False])

# Drop helper columns
df_pm_view = df_pm_view.drop(columns=["CategorySort", "DelayDays"])


# === Write to Excel ===
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    # PM View first
    df_pm_view.to_excel(writer, sheet_name="PM View", index=False)

    # Then full issue list
    df_issues.to_excel(writer, sheet_name="SF Issues", index=False)

    # Format PM View sheet
    ws_pm = writer.sheets["PM View"]
    end_col_pm = get_column_letter(ws_pm.max_column)
    end_row_pm = ws_pm.max_row
    table_range_pm = f"A1:{end_col_pm}{end_row_pm}"

    table_pm = Table(displayName="PMView", ref=table_range_pm)
    style_pm = TableStyleInfo(name="TableStyleLight1", showRowStripes=False, showColumnStripes=False)
    table_pm.tableStyleInfo = style_pm
    ws_pm.add_table(table_pm)

    for col in ws_pm.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws_pm.column_dimensions[col[0].column_letter].width = max_len + 2

    # Format SF Issues sheet
    ws_all = writer.sheets["SF Issues"]
    end_col_all = get_column_letter(ws_all.max_column)
    end_row_all = ws_all.max_row
    table_range_all = f"A1:{end_col_all}{end_row_all}"

    table_all = Table(displayName="SFIssues", ref=table_range_all)
    style_all = TableStyleInfo(name="TableStyleLight1", showRowStripes=False, showColumnStripes=False)
    table_all.tableStyleInfo = style_all
    ws_all.add_table(table_all)

    for col in ws_all.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws_all.column_dimensions[col[0].column_letter].width = max_len + 2

print(f"✅ Exported to {output_path}")
