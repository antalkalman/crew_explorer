import streamlit as st
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
import shutil

def color_status(val):
    colors = {
        "Prepped": "color: grey",
        "Approved": "color: green",
        "Signed": "color: blue",
        "Correct": "color: orange",
        "": "color: red"
    }
    return colors.get(val, "")

st.set_page_config(layout="wide")

# --- Load and Clean Data ---
if "df" not in st.session_state:
    base_paths = [
        "/Users/kalmanantal/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/visual_test/data_for_visual_test.xlsx",
        "/Users/antalkalman/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/visual_test/data_for_visual_test.xlsx"
    ]
    for path in base_paths:
        if os.path.exists(path):
            data_path = path
            break
    else:
        st.error("❌ Could not find data_for_visual_test.xlsx in expected locations.")
        st.stop()

    df_loaded = pd.read_excel(data_path)
    df_loaded.columns = df_loaded.columns.str.strip()
    df_loaded["Date"] = pd.to_datetime(df_loaded["Date"])
    valid_statuses = {"Prepped", "Approved", "Signed", "Correct"}
    df_loaded["Status"] = df_loaded["Status"].apply(lambda s: s if s in valid_statuses else "")
    st.session_state.df = df_loaded

df = st.session_state.df

# --- Sidebar Filters ---
st.sidebar.header("Filters")

unit_options = sorted(df["Unit"].dropna().unique())
dept_options = ["All"] + sorted(df["Department"].dropna().unique())
date_options = sorted(df["Date"].dt.date.unique())
id_options = sorted(df["ID"].dropna().unique())
status_options = ["Prepped", "Approved", "Signed", "Correct", ""]
status_labels = {
    "Prepped": "Prepped",
    "Approved": "Approved",
    "Signed": "Signed",
    "Correct": "Correct",
    "": "Empty"
}

status_filter_raw = st.sidebar.multiselect(
    "Filter by State",
    [status_labels[s] for s in status_options],
    default=[status_labels[s] for s in status_options]
)
label_to_status = {v: k for k, v in status_labels.items()}
status_filter = [label_to_status[label] for label in status_filter_raw]

unit_filter = st.sidebar.multiselect("Select Unit(s)", unit_options)
if not unit_filter:
    unit_filter = unit_options

date_filter = st.sidebar.multiselect("Select Date(s)", date_options)
if not date_filter:
    date_filter = date_options

id_filter = st.sidebar.multiselect("Select ID(s)", id_options)
if not id_filter:
    id_filter = id_options

# Department dropdown + navigation
if "current_dept_index" not in st.session_state:
    st.session_state.current_dept_index = 0

selected_dept = st.sidebar.selectbox("Select Department", dept_options)

if selected_dept != "All":
    if selected_dept in dept_options:
        st.session_state.current_dept_index = dept_options.index(selected_dept)

    col1, col2 = st.sidebar.columns([1, 1])
    with col1:
        if st.button("⬅️ Previous"):
            st.session_state.current_dept_index = max(1, st.session_state.current_dept_index - 1)
    with col2:
        if st.button("Next ➡️"):
            st.session_state.current_dept_index = min(len(dept_options) - 1, st.session_state.current_dept_index + 1)

    selected_dept = dept_options[st.session_state.current_dept_index]
    st.sidebar.markdown(f"### 📦 Department: `{selected_dept}`")
    department_filter = [selected_dept]
else:
    department_filter = sorted(df["Department"].dropna().unique())

# --- Filter Data ---
filtered_df = df[
    (df["Unit"].isin(unit_filter)) &
    (df["Department"].isin(department_filter)) &
    (df["Date"].dt.date.isin(date_filter)) &
    (df["ID"].isin(id_filter))
]
filtered_df = filtered_df[filtered_df["Status"].fillna("").isin(status_filter)]

# Optional: hide empty/zero sum
hide_zero = st.sidebar.checkbox("🔘 Hide rows with zero/empty Sum", value=False)
if hide_zero:
    filtered_df = filtered_df[filtered_df["Sum"].fillna(0) > 0]

# --- Totals
summary_totals = filtered_df[["OT", "Meal Penalty", "TA", "Sum"]].sum(numeric_only=True).round(0).astype(int)
st.markdown(
    f"### 🧮 Totals for current selection:\n"
    f"- **OT:** `{summary_totals['OT']:,}`  "
    f"- **MP:** `{summary_totals['Meal Penalty']:,}`  "
    f"- **TA:** `{summary_totals['TA']:,}`  "
    f"- **Sum:** `{summary_totals['Sum']:,}`"
)

# --- Set Status
st.markdown("### Set State for All Visible Rows")
status_labels_display = {
    "Prepped": "🟢 Prepped",
    "Approved": "🔵 Approved",
    "Signed": "🟣 Signed",
    "Correct": "🟠 Correct",
    "": "🔴 Empty"
}
status_reverse = {v: k for k, v in status_labels_display.items()}
new_status_label = st.radio(
    "👁️ Choose State to Apply:",
    options=list(status_labels_display.values()),
    horizontal=True
)
new_status = status_reverse[new_status_label]

if st.button("Apply State to All Visible"):
    ids_to_update = filtered_df["ID"].dropna().tolist()
    df.loc[df["ID"].isin(ids_to_update), "Status"] = new_status
    st.success("✅ State updated for visible rows.")

# --- Title & Display ---
st.title("OT Summary – Visual Test")

for dept in filtered_df["Department"].dropna().unique():
    dept_df = filtered_df[filtered_df["Department"] == dept]
    st.subheader(f"Department: {dept}")

    dept_df_display = dept_df.copy()
    dept_df_display["Date"] = dept_df_display["Date"].dt.strftime("%m/%d")
    dept_df_display["Start"] = dept_df_display["Start"].apply(lambda x: x.strftime("%H:%M") if pd.notnull(x) else "")
    dept_df_display["End"] = dept_df_display["End"].apply(lambda x: x.strftime("%H:%M") if pd.notnull(x) else "")
    dept_df_display["Sum"] = dept_df_display["Sum"].apply(lambda x: f"{int(round(x)):,}" if pd.notnull(x) else "")

    display_cols = ["Status", "ID", "Date", "Name", "Title", "Start", "End", "OT", "Meal Penalty", "TA", "Sum"]
    styled_df = dept_df_display[display_cols].style.applymap(color_status, subset=["Status"])
    styled_df = styled_df.format({
        "OT": lambda v: f"{int(v)}" if pd.notna(v) and float(v).is_integer() else f"{v:.2f}" if pd.notna(v) else "",
        "Meal Penalty": lambda v: f"{int(v)}" if pd.notna(v) and float(v).is_integer() else f"{v:.2f}" if pd.notna(v) else "",
        "TA": lambda v: f"{int(v)}" if pd.notna(v) and float(v).is_integer() else f"{v:.2f}" if pd.notna(v) else ""
    })

    st.dataframe(styled_df, use_container_width=True, hide_index=True)

    totals = dept_df[["OT", "Meal Penalty", "TA", "Sum"]].sum(numeric_only=True).round(0).astype(int)
    st.markdown(
        f"**Subtotal – OT:** `{totals['OT']:,}`  "
        f"**MP:** `{totals['Meal Penalty']:,}`  "
        f"**TA:** `{totals['TA']:,}`  "
        f"**Sum:** `{totals['Sum']:,}`"
    )

# --- Export ---
st.download_button(
    label="Download filtered data as CSV",
    data=filtered_df.to_csv(index=False),
    file_name="OT_summary_filtered.csv"
)

# --- Save Back ---
if st.button("💾 Save Changes to Excel"):
    now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    export_folders = [
        "/Users/kalmanantal/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/visual_test/",
        "/Users/antalkalman/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/visual_test/"
    ]
    for folder_path in export_folders:
        if os.path.exists(folder_path):
            folder = folder_path
            break
    else:
        st.error("❌ Could not find export folder.")
        st.stop()

    original_file = os.path.join(folder, "data_for_visual_test.xlsx")
    backup_file = os.path.join(folder, f"data_for_visual_test_BACKUP_{now}.xlsx")
    shutil.copy2(original_file, backup_file)

    wb = load_workbook(backup_file)
    ws = wb.active
    status_map = st.session_state.df.set_index("ID")["Status"].to_dict()
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    id_col = headers.index("ID") + 1
    status_col = headers.index("Status") + 1

    for row in ws.iter_rows(min_row=2):
        row_id = row[id_col - 1].value
        if row_id in status_map:
            row[status_col - 1].value = status_map[row_id]

    wb.save(backup_file)
    st.success(f"✅ Status column saved to: `{os.path.basename(backup_file)}`")
