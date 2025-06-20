import streamlit as st
import pandas as pd
import datetime
import os
from datetime import time, datetime
from openpyxl import load_workbook

def color_status(val):
    colors = {
        "Prepped": "color: grey",
        "Approved": "color: green",
        "Signed": "color: blue",
        "Correct": "color: orange",
        "": "color: red"
    }
    return colors.get(val, "")


# 👉 Enable wide mode
st.set_page_config(layout="wide")

# --- Load and Clean Data ---
if "df" not in st.session_state:
    # Cross-platform file path
    base_paths = [
        "/Users/kalmanantal/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/visual_test/data_for_visual_test.xlsx",
        "/Users/antalkalman/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/visual_test/data_for_visual_test.xlsx"
    ]

    for path in base_paths:
        if os.path.exists(path):
            data_path = path
            break
    else:
        st.error("❌ Could not find data_for_visual_test.xlsx in expected locations.")
        st.stop()

    df_loaded = pd.read_excel(data_path)

    df_loaded.columns = df_loaded.columns.str.strip()
    df_loaded["Date"] = pd.to_datetime(df_loaded["Date"])

    # ✅ Clean invalid statuses
    valid_statuses = {"Prepped", "Approved", "Signed", "Correct"}
    df_loaded["Status"] = df_loaded["Status"].apply(lambda s: s if s in valid_statuses else "")

    st.session_state.df = df_loaded

df = st.session_state.df

# --- Sidebar Filters ---
st.sidebar.header("Filters")

unit_options = sorted(df["Unit"].dropna().unique())
dept_options = sorted(df["Department"].dropna().unique())

# Track department index in session_state
if "dept_index" not in st.session_state:
    st.session_state.dept_index = 0

date_options = sorted(df["Date"].dt.date.unique())
id_options = sorted(df["ID"].dropna().unique())
status_options = ["Prepped", "Approved", "Signed", "Correct", ""]
status_labels = {
    "Prepped": "Prepped",
    "Approved": "Approved",
    "Signed": "Signed",
    "Correct": "Correct",
    "": "⛔ Empty"
}
status_filter_raw = st.sidebar.multiselect(
    "Filter by State",
    ["All"] + [status_labels[s] for s in status_options],
    default=["All"]
)

# Convert labels back to status values
if "All" in status_filter_raw:
    status_filter = status_options
else:
    label_to_status = {v: k for k, v in status_labels.items()}
    status_filter = [label_to_status[label] for label in status_filter_raw]

unit_filter = st.sidebar.multiselect("Select Unit(s)", ["All"] + unit_options, default=["All"])
if "All" in unit_filter or not unit_filter:
    unit_filter = unit_options

# Show only if 1 department selected
if "Next Department" not in st.session_state:
    st.session_state["Next Department"] = False

department_filter = st.sidebar.multiselect("Select Department(s)", ["All"] + dept_options, default=["All"])

# Handle "Next Department" button
if len(department_filter) == 1:
    if st.sidebar.button("➡️ Next Department"):
        current = dept_options.index(department_filter[0])
        next_index = (current + 1) % len(dept_options)
        department_filter = [dept_options[next_index]]
        st.session_state.dept_index = next_index


date_filter = st.sidebar.multiselect("Select Date(s)", ["All"] + date_options, default=["All"])
if "All" in date_filter or not date_filter:
    date_filter = date_options

id_filter = st.sidebar.multiselect("Select ID(s)", ["All"] + id_options, default=["All"])
if "All" in id_filter or not id_filter:
    id_filter = id_options

# --- Filter Main DataFrame ---
filtered_df = df[
    (df["Unit"].isin(unit_filter)) &
    (df["Department"].isin(department_filter)) &
    (df["Date"].dt.date.isin(date_filter)) &
    (df["ID"].isin(id_filter))
]

# Filter by status
if "All" not in status_filter:
    filtered_df = filtered_df[filtered_df["Status"].fillna("").isin(status_filter)]

# 🔘 Optional filter: Hide rows with zero or empty Sum
hide_zero = st.sidebar.checkbox("🔘 Hide rows with zero/empty Sum", value=False)
if hide_zero:
    filtered_df = filtered_df[filtered_df["Sum"].fillna(0) > 0]


# ✅ Show totals for the current visible selection
summary_totals = filtered_df[["OT", "Meal Penalty", "TA", "Sum"]].sum(numeric_only=True).round(0).astype(int)

st.markdown(
    f"### 🧮 Totals for current selection:\n"
    f"- **OT:** `{summary_totals['OT']:,}`  "
    f"- **MP:** `{summary_totals['Meal Penalty']:,}`  "
    f"- **TA:** `{summary_totals['TA']:,}`  "
    f"- **Sum:** `{summary_totals['Sum']:,}`"
)

# --- Set Status for All Visible ---
st.markdown("### Set State for All Visible Rows")
status_labels = {
    "Prepped": "🟢 Prepped",
    "Approved": "🔵 Approved",
    "Signed": "🟣 Signed",
    "Correct": "🟠 Correct",
    "": "🔴 Empty"
}
status_reverse = {v: k for k, v in status_labels.items()}  # for reverse lookup

new_status_label = st.radio(
    "👁️ Choose State to Apply:",
    options=list(status_labels.values()),
    horizontal=True
)
new_status = status_reverse[new_status_label]

if st.button("Apply State to All Visible"):
    ids_to_update = filtered_df["ID"].dropna().tolist()
    df.loc[df["ID"].isin(ids_to_update), "Status"] = new_status
    # refresh filtered_df to reflect changes
    filtered_df = df[
        (df["Unit"].isin(unit_filter)) &
        (df["Department"].isin(department_filter)) &
        (df["Date"].dt.date.isin(date_filter)) &
        (df["ID"].isin(id_filter))
    ]
    if "All" not in status_filter:
        filtered_df = filtered_df[filtered_df["Status"].isin(status_filter)]

# --- Title ---
st.title("OT Summary – Visual Test")

# --- Department Display with Scrollable Tables ---
for dept in filtered_df["Department"].dropna().unique():
    dept_df = filtered_df[filtered_df["Department"] == dept]
    st.subheader(f"Department: {dept}")

    # Format columns
    dept_df_display = dept_df.copy()
    dept_df_display["Date"] = dept_df_display["Date"].dt.strftime("%m/%d")
    dept_df_display["Start"] = dept_df_display["Start"].apply(lambda x: x.strftime("%H:%M") if pd.notnull(x) else "")
    dept_df_display["End"] = dept_df_display["End"].apply(lambda x: x.strftime("%H:%M") if pd.notnull(x) else "")
    dept_df_display["Sum"] = dept_df_display["Sum"].apply(lambda x: f"{int(round(x)):,}" if pd.notnull(x) else "")

    # Show table
    display_cols = ["Status", "ID", "Date", "Name", "Title", "Start", "End", "OT", "Meal Penalty", "TA", "Sum"]
    styled_df = dept_df_display[display_cols].style.applymap(color_status, subset=["Status"])


    def clean_number(val):
        if pd.isna(val):
            return ""
        if float(val).is_integer():
            return f"{int(val)}"
        return f"{val:.2f}"


    styled_df = styled_df.format({
        "OT": clean_number,
        "Meal Penalty": clean_number,
        "TA": clean_number
    })

    st.dataframe(styled_df, use_container_width=True, hide_index=True)

    # Subtotals
    totals = dept_df[["OT", "Meal Penalty", "TA", "Sum"]].sum(numeric_only=True).round(0).astype(int)
    st.markdown(
        f"**Subtotal – OT:** `{totals['OT']:,}`  "
        f"**MP:** `{totals['Meal Penalty']:,}`  "
        f"**TA:** `{totals['TA']:,}`  "
        f"**Sum:** `{totals['Sum']:,}`"
    )

# --- Export CSV with Current Status ---
st.download_button(
    label="Download filtered data as CSV",
    data=filtered_df.to_csv(index=False),
    file_name="OT_summary_filtered.csv"
)

# --- Save Back to Timestamped Excel File ---
from openpyxl import load_workbook

if st.button("💾 Save Changes to Excel"):
    now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    # Cross-platform export folder
    export_folders = [
        "/Users/kalmanantal/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/visual_test/",
        "/Users/antalkalman/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/visual_test/"
    ]

    for folder_path in export_folders:
        if os.path.exists(folder_path):
            folder = folder_path
            break
    else:
        st.error("❌ Could not find export folder.")
        st.stop()

    #filename = f"data_for_visual_test_BACKUP_{now}.xlsx"
    #full_path = os.path.join(folder, filename)

    original_file = os.path.join(folder, "data_for_visual_test.xlsx")
    backup_file = os.path.join(folder, f"data_for_visual_test_BACKUP_{now}.xlsx")

    # 🔁 Copy the original file first
    import shutil
    shutil.copy2(original_file, backup_file)

    # 📥 Load the backup
    wb = load_workbook(backup_file)
    ws = wb.active

    # Get status updates from session df
    status_map = st.session_state.df.set_index("ID")["Status"].to_dict()

    # Find column index for "ID" and "Status"
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    id_col = headers.index("ID") + 1
    status_col = headers.index("Status") + 1

    # Update each row
    for row in ws.iter_rows(min_row=2):
        row_id = row[id_col - 1].value
        if row_id in status_map:
            row[status_col - 1].value = status_map[row_id]

    wb.save(backup_file)
    st.success(f"✅ Status column saved to: `{os.path.basename(backup_file)}`")

