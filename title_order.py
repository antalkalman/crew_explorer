import os
import time
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# === CONFIG ===

# Base directory for output
desktop_path = "/Users/kalmanantal/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/API_test"
laptop_path = "/Users/antalkalman/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/API_test"
base_dir = desktop_path if os.path.exists(desktop_path) else laptop_path
os.makedirs(base_dir, exist_ok=True)

# API setup
token = "4|FtRaHbpkyP7dj6geLbqQiUc2WznEEqbRhvOD2xPO1cee3bfd"
base_url = "https://manager.crewcall.hu/api/"
headers = {"Authorization": f"Bearer {token}"}

# Output file
output_file = os.path.join(base_dir, "CrewManager_Project_Title_Sorted.xlsx")

# List of project names to include
included_projects = ["3BP", "BETAMAX", "Hidden Hand 2", "Oasis", "Seven Sisters"]

# === FUNCTION: Fetch data from API ===
def fetch_data(endpoint, retries=3, delay=5):
    for attempt in range(retries):
        try:
            response = requests.get(f"{base_url}{endpoint}", headers=headers, timeout=30)
            response.raise_for_status()
            return response.json().get("data", [])
        except requests.exceptions.RequestException as e:
            print(f"❌ {endpoint} – Attempt {attempt + 1} failed: {e}")
            if attempt < retries - 1:
                time.sleep(delay)
            else:
                raise

# === FETCH DATA ===
projects = fetch_data("project")
departments = fetch_data("department")
titles = fetch_data("job_title")

# === BUILD LOOKUPS ===
project_lookup = {p["id"]: p["name"] for p in projects}
dept_lookup = {d["id"]: {"name": d["name"], "sort": d["sort"]} for d in departments}

# Convert to DataFrames
df_title = pd.DataFrame(titles)
df_dept = pd.DataFrame(departments)

# === BUILD EXCEL OUTPUT ===
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    for pid, pname in project_lookup.items():
        if pname not in included_projects:
            continue

        title_filtered = df_title[df_title["project_id"] == pid][["name", "project_department_id", "sort"]].rename(
            columns={"name": "Title", "project_department_id": "Department ID", "sort": "Title Sort"}
        )
        title_filtered["Department"] = title_filtered["Department ID"].map(lambda x: dept_lookup.get(x, {}).get("name", ""))
        title_filtered["Department Sort"] = title_filtered["Department ID"].map(lambda x: dept_lookup.get(x, {}).get("sort", 0))

        result = title_filtered[["Department", "Title", "Department Sort", "Title Sort"]]
        result = result.sort_values(by=["Department Sort", "Title Sort"]).reset_index(drop=True)

        result.to_excel(writer, sheet_name=pname[:31], index=False)

# === ADD EXCEL TABLE FORMATTING ===
wb = load_workbook(output_file)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column
    last_col_letter = get_column_letter(max_col)
    table_ref = f"A1:{last_col_letter}{max_row}"
    table = Table(displayName=f"{sheet_name.replace(' ', '_')}_Table", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

wb.save(output_file)

print(f"✅ Excel export created with proper tables:\n{output_file}")
