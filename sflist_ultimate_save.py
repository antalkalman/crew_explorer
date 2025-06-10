# sflist_ultimate.py — The Master Morning Script
# This script compares the latest SF lists and builds a detailed Excel summary.

import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# === 📁 Set base folder for both machines ===
desktop_path = "/Users/kalmanantal/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive"
laptop_path = "/Users/antalkalman/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive"
base_dir = desktop_path if os.path.exists(desktop_path) else laptop_path

print("\n📌 Starting sflist_ultimate.py...")

# === 🧠 Helper: Normalize strings ===
def normalize(df):
    return df.apply(lambda col: col.str.strip().str.lower() if col.dtype == "object" else col)

# === 🧠 Helper: Detect issues (from sf_state_simple) ===
def detect_sf_issues(df_raw):
    import numpy as np

    excluded_projects = {"Death By Lightning", "Test", "Ballerina Overdrive", "Icebreaker"}

    df = df_raw.copy()
    df = df[~df["Project"].isin(excluded_projects)]
    df = df[df["State"] != "paused"]
    df = df[~df["Sf number"].astype(str).str.upper().str.startswith("CL")]

    def safe_str(x):
        return str(x).strip() if not pd.isna(x) else ""

    def clean_account(raw):
        digits = re.sub(r'\D', '', str(raw))
        return f"{digits[:8]}-{digits[8:16]}-{digits[16:24]}" if len(digits) >= 16 else digits

    def merge_tax_number(df):
        df["Personal: Tax number / Adóazonosító jel"] = df["Personal: Tax number / Adóazonosító jel"].fillna("").astype(str).str.strip()
        df["Company: VAT number / Adószám"] = df["Company: VAT number / Adószám"].fillna("").astype(str).str.strip()
        df["Tax Number"] = df["Personal: Tax number / Adóazonosító jel"]
        df["Tax Number"] = df["Tax Number"].where(df["Tax Number"] != "", df["Company: VAT number / Adószám"])
        df["Tax Number"] = df["Tax Number"].replace("", np.nan)
        return df

    def merge_bank_account(df):
        df["Bank Account"] = df["Personal: Bank account number / Bankszámlaszám"].apply(safe_str)
        df["Bank Account"] = df["Bank Account"].where(df["Bank Account"] != "", df["Company: Bank account number / Bankszámlaszám"].apply(safe_str))
        df["Bank Account"] = df["Bank Account"].where(df["Bank Account"] != "", df["Bank account number"].apply(safe_str))
        df["Bank Account"] = df["Bank Account"].apply(clean_account)
        return df

    def is_effectively_blank(s):
        if pd.isna(s): return True
        s = str(s).strip()
        return s == "" or s.lower() == "nan" or not re.search(r'[A-Za-z0-9]', s)

    def has_fee(row):
        return not (pd.isna(row.get("Daily fee")) and pd.isna(row.get("Weekly fee")))

    def find_issues(row, today):
        issues = []
        sf_type = str(row.get("Sf number", ""))[:2]

        if row.get("State") not in {"accepted", "signed"} and pd.notna(row.get("Start date")) and row["Start date"] < today:
            days_late = (today - row["Start date"]).days
            issues.append(f"Start date in past but not yet signed or accepted ({days_late} days)")

        if row.get("State") in {"accepted", "signed"}:
            required = [
                "Project department", "Project job title", "Project unit",
                "Surname", "Firstname", "Mobile number",
                "Crew list name", "Crew email", "Start date", "End date", "Deal type",
                "Project overtime", "Project turnaround", "Project working hour"
            ]

            if sf_type == "BD":
                required = [f for f in required if f not in {"Surname", "Firstname", "Mobile number", "Crew list name", "Crew email"}]

            for field in required:
                if is_effectively_blank(row.get(field, "")):
                    issues.append(field)

            if not has_fee(row):
                issues.append("Daily fee / Weekly fee")

            tax_number = row.get("Tax Number", "")
            company_name = row.get("Company: Company name / Cégnév", "")
            bank_account = row.get("Bank Account", "")

            if sf_type in {"BD", "DL"}:
                if is_effectively_blank(tax_number) and is_effectively_blank(company_name):
                    issues.append("Tax Number/Company")
            else:
                if is_effectively_blank(tax_number):
                    issues.append("Tax Number")
                if is_effectively_blank(bank_account):
                    issues.append("Bank Account")

        return ", ".join(issues)

    df["Start date"] = pd.to_datetime(df["Start date"], errors="coerce")
    today = pd.Timestamp(datetime.today().date())

    df = merge_tax_number(merge_bank_account(df))
    df["SF Key"] = df["Sf number"].astype(str) + " - " + df["Project"].astype(str)
    df["Issues"] = df.apply(lambda x: find_issues(x, today), axis=1)
    return df

# The rest of the code continues as-is...


# === 🧠 Helper: Get latest two SFlist CSVs ===
def get_latest_two_sflists(folder):
    pattern = re.compile(r"SFlist_(\d{8}_\d{4})\.csv$")
    sf_files = []
    for f in os.listdir(folder):
        match = pattern.match(f)
        if match:
            try:
                ts = datetime.strptime(match.group(1), "%Y%m%d_%H%M")
                sf_files.append((ts, f))
            except ValueError:
                continue
    if len(sf_files) < 2:
        raise FileNotFoundError("❌ Need at least two SFlist_*.csv files to compare.")
    sf_files.sort(reverse=True)
    return os.path.join(folder, sf_files[0][1]), os.path.join(folder, sf_files[1][1])

# === ✅ Step 1: Compare SFlist files ===
def compare_sflists(new_file_path, old_file_path):
    new_df = pd.read_csv(new_file_path, dtype=str).fillna("")
    old_df = pd.read_csv(old_file_path, dtype=str).fillna("")

    directions_path = os.path.join(base_dir, "compare_directions.xlsx")
    df_directions = pd.read_excel(directions_path)
    df_directions.columns = [col.strip() for col in df_directions.columns[:2]]
    field_to_category = dict(zip(
        df_directions.iloc[:, 0].astype(str).str.strip(),
        df_directions.iloc[:, 1].astype(str).str.strip()
    ))

    def normalize(df):
        return df.applymap(lambda x: x.strip().lower() if isinstance(x, str) else x)

    norm_new = normalize(new_df)
    norm_old = normalize(old_df)
    old_dict = norm_old.set_index("ID").to_dict(orient="index")

    def compare_row(row):
        row_id = row["ID"]
        if row_id not in old_dict:
            return "new"
        old_row = old_dict[row_id]
        diffs = []
        for col in norm_new.columns:
            if col == "ID" or field_to_category.get(col, "") in {"skip", "del"}:
                continue
            if row[col] != old_row.get(col, ""):
                diffs.append(col)
        return ", ".join(diffs) if diffs else "unchanged"

    norm_new["Change Status"] = norm_new.apply(compare_row, axis=1)
    output = new_df.copy()
    output["Change Status"] = norm_new["Change Status"]

    def get_primary_category(change_str):
        if change_str in ("", "unchanged", "new"):
            return change_str
        first_field = change_str.split(",")[0].strip()
        return field_to_category.get(first_field, "unknown")

    output["Category"] = output["Change Status"].apply(get_primary_category)

    # Remove fields marked as 'del'
    del_fields = [field for field, cat in field_to_category.items() if cat == "del"]
    keep_cols = [col for col in output.columns if col not in del_fields]
    output = output[keep_cols]

    return output, new_df

# === ✅ Step 2: Build Daily_processed_SFs.xlsx ===
# === ✅ Step 2: Build Daily_processed_SFs.xlsx ===
def save_final_excel(full_df, comparison_df):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = os.path.join(base_dir, f"Daily_processed_SFs_{timestamp}.xlsx")

    # Keep only rows with actual changes
    comparison_filtered = comparison_df[comparison_df["Change Status"] != "unchanged"].copy()

    # Reorder columns in Comparison tab
    cols = comparison_filtered.columns.tolist()
    if "Category" in cols and "Change Status" in cols:
        cols.remove("Category")
        cols.remove("Change Status")
        cols = ["Category", "Change Status"] + cols
        comparison_filtered = comparison_filtered[cols]

    # Detect issues in Full SFlist
    issues_df = detect_sf_issues(full_df)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # 🟢 Tab 1 — Full SFlist
        full_df.to_excel(writer, sheet_name="Full SFlist", index=False)
        # 🟡 Tab 2 — Comparison (filtered)
        comparison_filtered.to_excel(writer, sheet_name="Comparison", index=False)
        # 🔴 Tab 3 — All Issues
        issues_df.to_excel(writer, sheet_name="All Issues", index=False)

    # === Format Excel tables ===
    wb = load_workbook(out_path)
    for i, sheet in enumerate(["Full SFlist", "Comparison", "All Issues"], start=1):
        ws = wb[sheet]
        if ws.max_row <= 1:
            continue
        end_col = get_column_letter(ws.max_column)
        end_row = ws.max_row
        table_range = f"A1:{end_col}{end_row}"
        table = Table(displayName=f"Table{i}", ref=table_range)
        style = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
        table.tableStyleInfo = style
        ws.add_table(table)

        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Change Status":
                ws.column_dimensions[col_letter].width = 30  # Limit width
            else:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(out_path)
    print(f"✅ Excel saved to:\n{out_path}")



# === 🚀 MAIN EXECUTION ===
if __name__ == "__main__":
    print("📌 Starting sflist_ultimate.py...")

    # STEP 0: Locate input files
    new_file, old_file = get_latest_two_sflists(base_dir)
    print(f"🆕 Newest: {os.path.basename(new_file)}")
    print(f"📁 Older:  {os.path.basename(old_file)}")

    # STEP 1: Compare
    comparison_df, full_df = compare_sflists(new_file, old_file)

    # STEP 2: Create Excel
    save_final_excel(full_df, comparison_df)

    # === TODO ===
    # - STEP 3: Add "New Issues" tab (from compare_issues logic)
    # - STEP 4: Add "Resolved Issues" tab (from compare_issues logic)
    # - STEP 5: Add optional summary stats / color-coded analysis
