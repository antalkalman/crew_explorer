import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# === Set folder where files to compare are placed ===
desktop_compare = "/Users/kalmanantal/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive/Compare"
laptop_compare = "/Users/antalkalman/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive/Compare"
compare_dir = desktop_compare if os.path.exists(desktop_compare) else laptop_compare

# === Match SFlist_YYYYMMDD_HHMM.csv filenames ===
pattern = re.compile(r"SFlist_(\d{8}_\d{4})\.csv$")
sf_files = []

for f in os.listdir(compare_dir):
    match = pattern.match(f)
    if match:
        try:
            ts = datetime.strptime(match.group(1), "%Y%m%d_%H%M")
            sf_files.append((ts, f))
        except ValueError:
            continue

# Must have exactly two files
if len(sf_files) < 2:
    raise FileNotFoundError("❌ Need at least two SFlist_YYYYMMDD_HHMM.csv files in the 'Compare' folder.")

# Sort descending by timestamp
sf_files.sort(reverse=True)
latest_file = sf_files[0][1]
older_file = sf_files[1][1]

print(f"🔍 Comparing:\n  NEW : {latest_file}\n  OLD : {older_file}")

# === Load CSVs ===
sf_list = pd.read_csv(os.path.join(compare_dir, latest_file), dtype=str).fillna("")
af_list_old = pd.read_csv(os.path.join(compare_dir, older_file), dtype=str).fillna("")

# === Normalize data ===
def normalize(df):
    return df.applymap(lambda x: x.strip().lower() if isinstance(x, str) else x)

sf_clean = normalize(sf_list)
af_clean = normalize(af_list_old)
af_dict = af_clean.set_index("ID").to_dict(orient="index")

# === Compare rows ===
# === Compare rows, skipping renamed or ignored fields ===
def compare_row(row):
    row_id = row["ID"]
    if row_id not in af_dict:
        return "new"
    old_row = af_dict[row_id]
    skip_fields = {
        "ID", "Project startform", "Startform template",
        "business_type", "is_internal", "sort_order",
        "created_at", "updated_at", "downloaded_at", "Unit Sort",
        "Next of kin", "Contact number", "Dept Sort", "Title Sort", "Possible days"
    }  # fields to ignore
    diffs = [
        col for col in sf_clean.columns
        if col not in skip_fields and row[col] != old_row.get(col, "")
    ]
    return ", ".join(diffs) if diffs else "unchanged"


sf_clean["Change Status"] = sf_clean.apply(compare_row, axis=1)

# Combine with original data
output = sf_list.copy()
output["Change Status"] = sf_clean["Change Status"]

# === Final Excel export only ===
timestamp = datetime.now().strftime("%Y%m%d_%H%M")
excel_filename = f"SFlist_comparison_output_{timestamp}.xlsx"
excel_path = os.path.join(compare_dir, excel_filename)

with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    # Sheet 1: Comparison with Change Status
    output.to_excel(writer, sheet_name="Comparison", index=False)

    # Sheet 2: New file raw data
    sf_list.to_excel(writer, sheet_name="New SFlist", index=False)

    # Sheet 3: Old file raw data
    af_list_old.to_excel(writer, sheet_name="Old SFlist", index=False)

# === Apply Excel table formatting to 'Comparison' sheet only ===
wb = load_workbook(excel_path)
ws = wb["Comparison"]
end_col = get_column_letter(ws.max_column)
end_row = ws.max_row
table_range = f"A1:{end_col}{end_row}"

table = Table(displayName="ComparisonResult", ref=table_range)
style = TableStyleInfo(
    name="TableStyleLight1",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=False,
    showColumnStripes=False
)
table.tableStyleInfo = style
ws.add_table(table)

"""
# Optional: Adjust column widths
for col in ws.columns:
    max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_len + 2
"""

wb.save(excel_path)
print(f"📘 Excel with full details saved to:\n{excel_path}")