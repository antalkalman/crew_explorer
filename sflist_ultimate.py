import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# === 📁 Set base folder for both machines ===
desktop_path = "/Users/kalmanantal/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive"
laptop_path = "/Users/antalkalman/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/SF_Archive"
base_dir = desktop_path if os.path.exists(desktop_path) else laptop_path

print("\n📌 Starting sflist_ultimate.py...")

# === 🧠 Helper: Normalize strings ===
def normalize(df):
    return df.apply(lambda col: col.str.strip().str.lower() if col.dtype == "object" else col)

# === 🧠 Helper: Detect issues (from sf_state_simple) ===
def detect_sf_issues(df_raw):
    import numpy as np

    excluded_projects = {"Death By Lightning", "Test", "Ballerina Overdrive", "Icebreaker"}

    df = df_raw.copy()
    df = df[~df["Project"].isin(excluded_projects)]
    df = df[df["State"] != "paused"]
    df = df[~df["Sf number"].astype(str).str.upper().str.startswith("CL")]

    def safe_str(x):
        return str(x).strip() if not pd.isna(x) else ""

    def clean_account(raw):
        digits = re.sub(r'\D', '', str(raw))
        return f"{digits[:8]}-{digits[8:16]}-{digits[16:24]}" if len(digits) >= 16 else digits

    def merge_tax_number(df):
        df["Personal: Tax number / Adóazonosító jel"] = df["Personal: Tax number / Adóazonosító jel"].fillna("").astype(str).str.strip()
        df["Company: VAT number / Adószám"] = df["Company: VAT number / Adószám"].fillna("").astype(str).str.strip()
        df["Tax Number"] = df["Personal: Tax number / Adóazonosító jel"]
        df["Tax Number"] = df["Tax Number"].where(df["Tax Number"] != "", df["Company: VAT number / Adószám"])
        df["Tax Number"] = df["Tax Number"].replace("", np.nan)
        return df

    def merge_bank_account(df):
        df["Bank Account"] = df["Personal: Bank account number / Bankszámlaszám"].apply(safe_str)
        df["Bank Account"] = df["Bank Account"].where(df["Bank Account"] != "", df["Company: Bank account number / Bankszámlaszám"].apply(safe_str))
        df["Bank Account"] = df["Bank Account"].where(df["Bank Account"] != "", df["Bank account number"].apply(safe_str))
        df["Bank Account"] = df["Bank Account"].apply(clean_account)
        return df

    def is_effectively_blank(s):
        if pd.isna(s): return True
        s = str(s).strip()
        return s == "" or s.lower() == "nan" or not re.search(r'[A-Za-z0-9]', s)

    def has_fee(row):
        return not (pd.isna(row.get("Daily fee")) and pd.isna(row.get("Weekly fee")))

    def find_issues(row, today):
        issues = []
        sf_type = str(row.get("Sf number", ""))[:2]

        if row.get("State") not in {"accepted", "signed"} and pd.notna(row.get("Start date")) and row["Start date"] < today:
            days_late = (today - row["Start date"]).days
            issues.append(f"Start date in past but not yet signed or accepted ({days_late} days)")

        if row.get("State") in {"accepted", "signed"}:
            required = [
                "Project department", "Project job title", "Project unit",
                "Surname", "Firstname", "Mobile number",
                "Crew list name", "Crew email", "Start date", "End date", "Deal type",
                "Project overtime", "Project turnaround", "Project working hour"
            ]

            if sf_type == "BD":
                required = [f for f in required if f not in {"Surname", "Firstname", "Mobile number", "Crew list name", "Crew email"}]

            for field in required:
                if is_effectively_blank(row.get(field, "")):
                    issues.append(field)

            if not has_fee(row):
                issues.append("Daily fee / Weekly fee")

            tax_number = row.get("Tax Number", "")
            company_name = row.get("Company: Company name / Cégnév", "")
            bank_account = row.get("Bank Account", "")

            if sf_type in {"BD", "DL"}:
                if is_effectively_blank(tax_number) and is_effectively_blank(company_name):
                    issues.append("Tax Number/Company")
            else:
                if is_effectively_blank(tax_number):
                    issues.append("Tax Number")
                if is_effectively_blank(bank_account):
                    issues.append("Bank Account")

        return ", ".join(issues)

    df["Start date"] = pd.to_datetime(df["Start date"], errors="coerce")
    today = pd.Timestamp(datetime.today().date())

    df = merge_tax_number(merge_bank_account(df))
    df["SF Key"] = df["Sf number"].astype(str) + " - " + df["Project"].astype(str)
    df["Issues"] = df.apply(lambda x: find_issues(x, today), axis=1)
    return df

# === 🚀 MAIN EXECUTION CONTINUES HERE ===

# STEP 0: Identify latest two SFlists
all_sflist_files = sorted([f for f in os.listdir(base_dir) if f.startswith("SFlist_") and f.endswith(".csv")])
latest_file = os.path.join(base_dir, all_sflist_files[-1])
previous_file = os.path.join(base_dir, all_sflist_files[-2])

# STEP 1: Load and compare
new_df = pd.read_csv(latest_file)
old_df = pd.read_csv(previous_file)

# Detect issues
new_issues = detect_sf_issues(new_df)
old_issues = detect_sf_issues(old_df)

# Filter to ones with issues
all_issues = new_issues[new_issues["Issues"] != ""]
prev_issues = old_issues[old_issues["Issues"] != ""]

# Match by SF Key
prev_issues_dict = dict(zip(prev_issues["SF Key"], prev_issues["Issues"]))

def classify_issue(row):
    sfkey = row["SF Key"]
    old = prev_issues_dict.get(sfkey, "")
    new = row["Issues"]
    if old == new:
        return "Unchanged"
    elif not old:
        return "New"
    elif not new:
        return "Resolved"
    return "Changed"

all_issues["Issue Status"] = all_issues.apply(classify_issue, axis=1)

# Derive tabs
new_issues_only = all_issues[all_issues["Issue Status"] == "New"].copy()
resolved_issues = prev_issues[~prev_issues["SF Key"].isin(all_issues["SF Key"])]

# STEP 2: Append to Excel file
output_filename = sorted([f for f in os.listdir(base_dir) if f.startswith("Daily_processed_SFs_") and f.endswith(".xlsx")])[-1]
output_path = os.path.join(base_dir, output_filename)

wb = load_workbook(output_path)

# Write each new tab
with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    all_issues.to_excel(writer, sheet_name="All Issues", index=False)
    new_issues_only.to_excel(writer, sheet_name="New Issues", index=False)
    resolved_issues.to_excel(writer, sheet_name="Resolved Issues", index=False)

print(f"✅ Added 3 issue tabs to:", output_path)
