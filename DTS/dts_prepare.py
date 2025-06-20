import pandas as pd
from rapidfuzz import fuzz
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# === Paths ===
from pathlib import Path

# Fixed base path
base_path = Path("/Users/kalmanantal/Dropbox (Personal)/WORK/Pioneer Pictures/Adatbázis/DTS test")

# Daily TS folder inside DTS test
dts_folder = base_path / "Daily TS"

# Output Excel file inside DTS test
output_path = base_path / "Corrected_Combined_Daily_Timesheets.xlsx"

# Latest SFlist CSV from DTS test/SF_Archive
sf_path = max(base_path.glob("SFlist_*.csv"), key=lambda f: f.stat().st_mtime)


print(f"📂 Using SFlist: {sf_path.name}")



print(f"📂 Using SFlist: {sf_path.name}")

# === Combine all DTS Excel files ===
combined_dts = []
for file in sorted(dts_folder.glob("*.xlsx")):
    try:
        df = pd.read_excel(file, skiprows=7, usecols="A:M")
        df.columns = df.columns.map(str)
        df["Source File"] = file.name
        combined_dts.append(df)
        print(f"🔄 Loaded {file.name}")
    except Exception as e:
        print(f"⚠️ Could not load {file.name}: {e}")

df_dts_all = pd.concat(combined_dts, ignore_index=True)

# === Drop invalid rows and keep dropped ones separately ===
df_dts_all.columns = df_dts_all.columns.map(str)
dropped_rows = df_dts_all[
    (df_dts_all["Beosztás"].isna())
    #(df_dts_all["TS státusz"].str.strip().str.lower() != "ok")
]
df_dts = df_dts_all.drop(dropped_rows.index).copy()

# === Load SF list ===
df_sf = pd.read_csv(sf_path)
assert all(col in df_sf.columns for col in ["Crew list name", "Project job title", "Sf number"])
sf_pairs = list(zip(df_sf["Crew list name"], df_sf["Project job title"], df_sf["Sf number"]))

# === Matching Helpers ===
def normalize(text):
    return str(text).strip().lower()

def find_best_match(name, title):
    name_norm = normalize(name)
    title_norm = normalize(title)
    best_score = 0
    best_match = None
    for sf_name, sf_title, sf_number in sf_pairs:
        name_score = fuzz.token_set_ratio(name_norm, normalize(sf_name))
        title_score = fuzz.token_sort_ratio(title_norm, normalize(sf_title))
        score = 0.7 * name_score + 0.3 * title_score
        if score > best_score:
            best_score = score
            best_match = (sf_name, sf_title, sf_number)
    return best_match if best_score >= 85 else None

def match_blank_deal_title(title):
    best_score = 0
    best_title = None
    best_sf_number = None
    df_bd = df_sf[df_sf["Sf number"].astype(str).str.startswith("BD")]
    for _, row in df_bd.iterrows():
        sf_title = row["Project job title"]
        score = fuzz.token_sort_ratio(normalize(title), normalize(sf_title))
        if score > best_score:
            best_score = score
            best_title = sf_title
            best_sf_number = row["Sf number"]
    if best_score >= 85:
        return best_title, best_sf_number
    return None, None

# === Apply Matching ===
corrected_names = []
corrected_titles = []
original_names = []
original_titles = []
match_flags = []
matched_sf_numbers = []

for _, row in df_dts.iterrows():
    orig_name = row["Név (angolul)"]
    orig_title = row["Beosztás"]
    original_names.append(orig_name)
    original_titles.append(orig_title)

    match = find_best_match(orig_name, orig_title)
    if match:
        matched_name, matched_title, sf_number = match
        corrected_names.append(matched_name)
        corrected_titles.append(matched_title)
        matched_sf_numbers.append(sf_number)
        match_flags.append("Changed" if (matched_name != orig_name or matched_title != orig_title) else "Same")
    else:
        bd_title, bd_sf_number = match_blank_deal_title(orig_title)
        if bd_title:
            corrected_names.append(orig_name)
            corrected_titles.append(bd_title)
            matched_sf_numbers.append(bd_sf_number)
            match_flags.append("Blank Deal")
        else:
            corrected_names.append(orig_name)
            corrected_titles.append(orig_title)
            matched_sf_numbers.append("")
            match_flags.append("Same")

# === Build final output DataFrame ===
df_result = df_dts.copy()
df_result["Original Name"] = original_names
df_result["Original Title"] = original_titles
df_result["Név (angolul)"] = corrected_names
df_result["Beosztás"] = corrected_titles
df_result["Match result"] = match_flags
df_result["Matched SF number"] = matched_sf_numbers

# === Write Excel with two sheets ===
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df_result.to_excel(writer, index=False, sheet_name="Corrected TS")
    dropped_rows.to_excel(writer, index=False, sheet_name="Dropped Rows")

# === Format Excel Table ===
wb = load_workbook(output_path)
ws = wb["Corrected TS"]
max_row = ws.max_row
max_col = ws.max_column
last_col = get_column_letter(max_col)

table = Table(displayName="CorrectedTable", ref=f"A1:{last_col}{max_row}")
table.tableStyleInfo = TableStyleInfo(
    name="TableStyleLight1",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=False,
    showColumnStripes=False
)
ws.add_table(table)

# Auto-adjust column widths
for col in ws.columns:
    max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)

wb.save(output_path)
print(f"\n✅ All done! File saved to: {output_path.name}")
